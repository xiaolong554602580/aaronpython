#-*-coding:utf-8-*-
'''
Json文件-->sqlite數據庫-->篩選後導出Excel

2020.2.9:
完善數據導入及導出

下一步計劃：
1.加入導出數據的字體由簡轉繁；
2.專利類別導出時轉為中文定義
'''
import openpyxl
import json,os,time
import sqlite3
#import zhtools


def IsExistenceDatabase(*BaitenData):
	'''當前路徑標準數據庫的建立'''
	conn = sqlite3.connect('Patents.db')
	c = conn.cursor()
	'''標準寫法，如果是已存在的表則不會創建，integer為	整形，text為字符型，not null為標識，意思非空即必須
	有值。unique為唯一不可重複。primary key主鍵'''
	c.execute('''CREATE TABLE IF NOT EXISTS PATENTS(
		id INT,pn TEXT,ti TEXT,an TEXT,pd TEXT,ad TEXT,pa TEXT,in1 TEXT,ls1 TEXT,\
		ab TEXT,apn TEXT,apd TEXT,ic2 TEXT,ic1 TEXT,pr TEXT,aa TEXT,agc TEXT,agt TEXT,\
		cty TEXT,ls1_2 TEXT,ls2_1 TEXT,cpa TEXT,type TEXT,lsnt TEXT,lsn2 TEXT,lsn1 TEXT,\
		lset TEXT,lse TEXT,ain TEXT,co TEXT,ac TEXT,PRIMARY KEY(id));''')
	#寫入數據庫
	for jsonData in BaitenData:
			id = jsonData['id']
			pn = jsonData['pn']
			ti = jsonData['ti']
			an = jsonData['an']
			pd = jsonData['pd']
			ad = jsonData['ad']
			pa = jsonData['pa']
			in1 = jsonData['in']
			ls1 = jsonData['ls1']
			ab = jsonData['ab']
			apn = jsonData['apn']
			apd = jsonData['apd']
			ic2 = jsonData['ic2']
			ic1 = jsonData['ic1']
			pr = jsonData['pr']
			aa = jsonData['aa']
			agc = jsonData['agc']
			agt = jsonData['agt']
			cty = jsonData['cty']
			ls1_2 = jsonData['ls1_2']
			ls2_1 = jsonData['ls2_1']
			cpa = jsonData['cpa']
			type = jsonData['type']
			lsnt = jsonData['lsnt']
			lsn2 = jsonData['lsn2']
			lsn1 = jsonData['lsn1']
			lset = jsonData['lset']
			lse = jsonData['lse']
			ain = jsonData['ain']
			co = jsonData['co']
			ac = jsonData['ac']
	sqlText = (str(id),str(pn),str(ti),str(an),str(pd),str(ad),str(pa),str(in1),str(ls1),str(ab),str(apn),str(apd),str(ic2),str(ic1),\
				str(pr),str(aa),str(agc),str(agt),str(cty),str(ls1_2),str(ls2_1),str(cpa),str(type),str(lsnt),str(lsn2),str(lsn1),str(lset),\
				str(lse),str(ain),str(co),str(ac))
	#INSERT OR IGNORE/REPLACE
	c.execute("INSERT OR REPLACE INTO PATENTS VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",sqlText)
	conn.commit()
	conn.close()

def openJsonFile(fileName):
	'''Json文件處理'''
	panTypes = ['id','pn','ti','an','pd','ad','pa','in','ls1','ab','apn','apd','ic2','ic1','pr','aa','agc','agt','cty',\
			'ls1_2','ls2_1','cpa','type','lsnt','lsn2','lsn1','lset','lse','ain','co','ac']
	try:
		f = open(fileName,"r",encoding="utf-8-sig")
		jsonData = json.load(f)
	except:
		print("文件格式錯誤")
		pass
	else:
		jsonDatas = jsonData['cubePatentSearchResponse']['documents']
		print("該文件讀取到{}條信息".format(len(jsonDatas)))
		i = 0 
		while i < len(jsonDatas):
			jsonData = jsonDatas[i]['field_values']
			for j in range(0,len(panTypes)):
				jsonData.setdefault(panTypes[j]) 
			#print(jsonData)
			IsExistenceDatabase(jsonData)
			i+=1
		f.close()
		print("寫入成功")

def JsonFileDir(filesPath):
	'''得到JSON文件目錄'''
	JsonFiles = []
	Files = os.listdir(filesPath)
	for JsonFile in Files:
		if os.path.splitext(JsonFile)[1] == ".json":
			JsonFiles.append(filesPath + "/" + JsonFile)
			print("獲取了{}文件；".format(JsonFile))
	print("總共得到了{}個文件。".format(len(JsonFiles)))
	time.sleep(3)
	return JsonFiles		

def main():
	'''批量寫入文件到數據庫'''
	jsonFileDir = JsonFileDir(r"d:\aaronmo\Desktop\日常使用\1-我的學習\PythonScript")#此處填寫Json文件夾的路徑
	filecount = len(jsonFileDir)
	for fileName in jsonFileDir:
		openJsonFile(fileName)
		filecount-=1
		print('還有{}個文件。'.format(filecount))

def ExportDataToExcel():
	conn = sqlite3.connect('Patents.db')
	c = conn.cursor()
	c.execute('''SELECT type,ti,ab,pn,pd,pa,lsnt FROM PATENTS''')
	patents = c.fetchall()
	wb = openpyxl.Workbook()
	ws = wb.active
	headers = ['專利類型','專利名稱','文摘','公開號','公開日','權利人','主法狀']
	#ws['A1']=headers[0]
	ws.append(headers)
	i = 2
	for patent in patents:
		ws.cell(row=i,column=1,value=str(patent[0]))
		ws.cell(row=i,column=2,value=str(patent[1]))
		ws.cell(row=i,column=3,value=str(patent[2]))
		ws.cell(row=i,column=4,value=str(patent[3]))
		ws.cell(row=i,column=5,value=str(patent[4]))
		ws.cell(row=i,column=6,value=str(patent[5]))
		ws.cell(row=i,column=7,value=str(patent[6]))

		print(i)
		i += 1
	wb.save('PATENTS.xlsx')

#main()
ExportDataToExcel()
