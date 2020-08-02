#-*-coding:utf-8-*-

#這是一個從SQLITE寫入XLSX的腳本

import openpyxl
import sqlite3
import zhtools


class XlDataDown():
	"""docstring for XlDataDown"""
	def __init__(self):
		super(XlDataDown, self).__init__()
		c = self.open_database('z:/data/data.sqlite')
		c.execute('SELECT id,ad,type,ti,ab,pa,co,lsnt FROM print100')
		patents = c.fetchall()
		print('123')
		self.open_xlsx(patents)

		

	#Open database and get patents data
	def open_database(self,database):
		conn = sqlite3.connect(database)
		c = conn.cursor()
		return c
		#c.execute('''SELECT id,ad,type,ti,ab,pa,co,lsnt FROM''')

	#create excel for patent data
	def open_xlsx(self,patents):
		wb = openpyxl.Workbook()
		ws = wb.active
		#headers = ['專利申請號','申請日期','專利類型','專利名稱','摘要','權利人','地區','狀態','文件路徑']
		#for i in range(0,len(headers)):
			#ws.cell(row=1,column=i+1,value=headers[i])
		i = 1
		for patent in patents:
			
			ws.cell(row=i,column=1,value=patent[0])
			ws.cell(row=i,column=2,value=str(patent[1]))
			ws.cell(row=i,column=3,value=str(patent[2]))
			ws.cell(row=i,column=4,value=str(patent[3]))
			ws.cell(row=i,column=5,value=str(patent[4]))
			ws.cell(row=i,column=6,value=str(patent[5]))
			ws.cell(row=i,column=7,value=str(patent[6]))
			ws.cell(row=i,column=8,value=str(patent[7]))
			print(i)
			i += 1

			
		wb.save('print100.xlsx')



if __name__ == '__main__':
	XlDataDown()
	